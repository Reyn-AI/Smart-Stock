from typing import Dict, List
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
import os
from operator import itemgetter
from copy import deepcopy
from collections import defaultdict
from src.utils.common import scientific_number_convert


def str2float(string: str):
    """字符串转float"""
    try:
        float(string)
        return float(string)
    except ValueError:
        return str(string)


class ExcelUtils:
    """可视化到excel"""

    def __init__(self, save_path="./stock_data.xlsx"):

        self.save_path = save_path
        os.makedirs(os.path.dirname(self.save_path), exist_ok=True)
        self._init_vars()

    def _init_vars(self):

        self.wb = Workbook()
        self.sheet_cout = 0

    def _init_head(self, sheet_name, head_names, row_id=1) -> Worksheet:

        sheet = self.wb.create_sheet(sheet_name, self.sheet_cout)
        self.sheet_cout += 1
        start_pos = ord("A")
        for i, name in enumerate(head_names):
            sheet[f"{chr(start_pos+i)}{row_id}"] = name

        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 45
        sheet.column_dimensions["E"].width = 25
        sheet.column_dimensions["F"].width = 35
        return sheet

    def _group_by(self, infos: List[Dict], keys: tuple):
        keys_recoder = defaultdict(list)
        for info in infos:
            key_recoder = tuple([info[x] for x in keys])
            keys_recoder[key_recoder].append(info)

        for k, v in keys_recoder.items():
            yield k, v

    def write_stock_list_excel(self, infos: Dict[List[Dict]]):
        """json写excel
            infos: {'sheet_name':list(dict)}
        """
        for sheet_name, info in infos.items():
            self._init_sheet_by_dict(infos=info, sheet_name=sheet_name)
        return self

    def _init_sheet_by_dict(
        self,
        infos: List[Dict],
        sheet_name="op-infos",
        row_start=2,
        col_start=1,
    ):

        if len(infos) == 0:
            return
        sheet_name = f"{sheet_name}"
        head_names = infos[0].keys()
        sheet = self._init_head(
            sheet_name=sheet_name, head_names=head_names, row_id=row_start - 1
        )
        for i, info in enumerate(infos):
            for j, kv in enumerate(info.items()):
                k, v = kv
                sheet.cell(row=row_start + i, column=col_start + j, value=scientific_number_convert(v))

    def save(self):
        """保存excel"""
        del self.wb["Sheet"]
        self.wb.save(self.save_path)




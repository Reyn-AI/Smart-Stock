from typing import Dict, List
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re
import os
from operator import itemgetter
from copy import deepcopy
from collections import defaultdict
from src.utils.common import scientific_number_convert, get_logger
from src.view.base import BaseViewer


def str2float(string: str):
    """字符串转float"""
    try:
        float(string)
        return float(string)
    except ValueError:
        return str(string)


class BaseExcel(BaseViewer):
    """可视化到excel"""

    def __init__(self, save_path="./stock_data.xlsx"):

        self.save_path = save_path
        os.makedirs(os.path.dirname(self.save_path), exist_ok=True)
        self._init_vars()

    def _init_vars(self):

        self.wb = Workbook()
        self.sheet_cout = 0
        self.align = Alignment(horizontal='center', vertical='center',wrapText=True)

    def visual(self, infos: Dict, *args, **kwargs):
        """可视化抽象方法"""
        return super().visual(infos)

    def _init_head(self, sheet_name, head_names, row_id=1) -> Worksheet:
        sheet = self.wb.create_sheet(sheet_name, self.sheet_cout)
        self.sheet_cout += 1
        for i, name in enumerate(head_names, start=1):
            sheet.cell(row=row_id, column=i, value=name)
            sheet.cell(row_id, i).alignment = self.align
        sheet.freeze_panes = "C2"
        # sheet.column_dimensions["A"].width = 30
        # sheet.column_dimensions["B"].width = 15
        # sheet.column_dimensions["C"].width = 20
        # sheet.column_dimensions["D"].width = 45
        # sheet.column_dimensions["E"].width = 25
        # sheet.column_dimensions["F"].width = 35
        return self.adaptive_column_width(sheet=sheet)

    def adaptive_column_width(self, sheet):
        """自适应列宽"""
        # 设置一个字典用于保存列宽数据
        dims = {}

        # 遍历表格数据，获取自适应列宽数据
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                    # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                    # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                    cell_len = 0.7 * len(
                        re.findall("([\u4e00-\u9fa5])", str(cell.value))
                    ) + len(str(cell.value))
                    dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
        for col, value in dims.items():
            # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
            sheet.column_dimensions[get_column_letter(col)].width = value + 5
        return sheet

    def _group_by(self, infos: List[Dict], keys: tuple):
        keys_recoder = defaultdict(list)
        for info in infos:
            key_recoder = tuple([info[x] for x in keys])
            keys_recoder[key_recoder].append(info)

        for k, v in keys_recoder.items():
            yield k, v

    def write_stock_list_excel(self, infos: Dict):
        """json写excel
        infos: {'sheet_name':list(dict)}
        """
        for sheet_name, info in infos.items():
            if 'plot' in sheet_name:
                #跳过绘制的数据
                continue
            self._init_sheet_by_dict(infos=info, sheet_name=sheet_name)
        return self

    def _init_sheet_by_dict(
        self,
        infos: List[Dict],
        sheet_name="op-infos",
        row_start=2,
        col_start=1,
    ):

        if len(infos) == 0:
            return
        sheet_name = f"{sheet_name}"
        head_names = infos[0].keys()
        sheet = self._init_head(
            sheet_name=sheet_name, head_names=head_names, row_id=row_start - 1
        )
        for i, info in enumerate(infos):
            for j, kv in enumerate(info.items()):
                _, v = kv
                sheet.cell(
                    row=row_start + i,
                    column=col_start + j,
                    value=scientific_number_convert(v),
                )
                sheet.cell(row_start + i, col_start + j).alignment = self.align

    def save(self):
        """保存excel"""
        del self.wb["Sheet"]
        self.wb.save(self.save_path)
        get_logger().info("Finished, Excel save in %s", self.save_path)


class WZWExcel(BaseExcel):

    def visual(self, infos: Dict, *args, **kwargs):
        self.write_stock_list_excel(infos).save()

class TuShareExcel(BaseExcel):
    def visual(self, infos: Dict, *args, **kwargs):
        self.write_stock_list_excel(infos).save()

class AkShareExcel(BaseExcel):
    def visual(self, infos: Dict, *args, **kwargs):
        self.write_stock_list_excel(infos).save()